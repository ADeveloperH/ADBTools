#!/usr/bin/env python3
from __future__ import annotations

import json
import shutil
import sys
import tempfile
import zipfile
from dataclasses import dataclass, asdict
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook


@dataclass
class Project:
    id: str
    project_name: str
    app_name: str
    package: str
    store_name: str = ""
    company_name: str = ""


def find_archive(cwd: Path) -> Path:
    archives = sorted(p for p in cwd.iterdir() if p.is_file() and p.suffix.lower() == ".zip")
    if not archives:
        raise FileNotFoundError("No .zip archive found in current directory.")
    return archives[0]


def extract_xlsx(archive: Path) -> tuple[Path, Path]:
    tmpdir = Path(tempfile.mkdtemp(prefix="adb_project_import_"))
    with zipfile.ZipFile(archive) as zf:
        workbook_name = next((name for name in zf.namelist() if name.lower().endswith(".xlsx")), None)
        if not workbook_name:
            raise FileNotFoundError("No .xlsx workbook found in archive.")
        out = tmpdir / Path(workbook_name).name
        with zf.open(workbook_name) as src, out.open("wb") as dst:
            shutil.copyfileobj(src, dst)
    return tmpdir, out


def normalize(value) -> str:
    return "" if value is None else str(value).strip()


def read_projects_from_workbook(workbook: Path) -> tuple[list[Project], int]:
    wb = load_workbook(workbook, read_only=True, data_only=True)
    if "APP 信息总表" not in wb.sheetnames:
        raise ValueError("Workbook missing 'APP 信息总表' sheet.")
    ws = wb["APP 信息总表"]
    rows = ws.iter_rows(values_only=True)
    header = [normalize(v) for v in next(rows, ())]
    if not header or header[0] != "AppName":
        raise ValueError("Unexpected summary sheet layout.")

    result: list[Project] = []
    skipped = 0
    for row in rows:
        values = [normalize(v) for v in row]
        if not any(values):
            skipped += 1
            continue
        app_name = values[0] if len(values) > 0 else ""
        store_name = values[1] if len(values) > 1 else ""
        company_name = values[2] if len(values) > 2 else ""
        package = values[3] if len(values) > 3 else ""
        if not package:
            skipped += 1
            continue
        project_name = store_name or app_name
        project_id = package.replace(".", "_")
        result.append(
            Project(
                id=project_id,
                project_name=project_name,
                app_name=app_name,
                package=package,
                store_name=store_name,
                company_name=company_name,
            )
        )
    return result, skipped


def load_existing(projects_json: Path) -> list[Project]:
    if not projects_json.exists():
        return []
    data = json.loads(projects_json.read_text(encoding="utf-8"))
    return [Project(**item) for item in data.get("projects", [])]


def merge(existing: list[Project], incoming: list[Project]) -> tuple[list[Project], int, int]:
    by_package = {item.package: item for item in existing}
    updated = 0
    added = 0
    new_projects: list[Project] = []
    seen_incoming: set[str] = set()
    for project in incoming:
        if project.package in seen_incoming:
            continue
        seen_incoming.add(project.package)
        if project.package in by_package:
            updated += 1
        else:
            added += 1
            new_projects.append(project)
        by_package[project.package] = project
    # 保持已有项目的原顺序（仅更新字段），新增项目按导入顺序追加到末尾。
    merged = [by_package[item.package] for item in existing if item.package in by_package]
    merged += new_projects
    return merged, added, updated


def index_by_package(projects: Iterable[Project]) -> dict[str, Project]:
    return {project.package: project for project in projects}


def write_projects(projects_json: Path, projects: Iterable[Project]) -> None:
    payload = {"projects": [asdict(project) for project in projects]}
    projects_json.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def main() -> int:
    cwd = Path.cwd()
    archive = find_archive(cwd)
    tmpdir, workbook = extract_xlsx(archive)
    try:
        incoming, skipped = read_projects_from_workbook(workbook)
        projects_json = cwd / "config" / "projects.json"
        existing = load_existing(projects_json)
        existing_by_package = index_by_package(existing)
        merged, added, updated = merge(existing, incoming)
        write_projects(projects_json, merged)
        added_items = [project for project in incoming if project.package not in existing_by_package]
        updated_items = [project for project in incoming if project.package in existing_by_package]
        changed_items = [
            project
            for project in updated_items
            if any(
                [
                    project.project_name != existing_by_package[project.package].project_name,
                    project.app_name != existing_by_package[project.package].app_name,
                    project.store_name != existing_by_package[project.package].store_name,
                    project.company_name != existing_by_package[project.package].company_name,
                ]
            )
        ]

        print(f"archive: {archive}")
        print(f"imported: {added}")
        print(f"updated: {updated}")
        print(f"skipped: {skipped}")
        if added_items:
            print("新增项目:")
            for project in added_items:
                print(f"- {project.project_name} / {project.app_name} / {project.package}")
        if updated_items:
            print("更新项目:")
            for project in updated_items:
                old = existing_by_package[project.package]
                if project.project_name == old.project_name and project.app_name == old.app_name and project.store_name == old.store_name and project.company_name == old.company_name:
                    print(f"- {project.project_name} / {project.app_name} / {project.package}（未变更）")
                else:
                    print(
                        f"- {project.project_name} / {project.app_name} / {project.package}"
                        f"（原: {old.project_name} / {old.app_name} / {old.package}）"
                    )
        if changed_items:
            print(f"实际发生字段变化的项目: {len(changed_items)}")
        else:
            print("实际发生字段变化的项目: 0")
        return 0
    finally:
        shutil.rmtree(tmpdir, ignore_errors=True)


if __name__ == "__main__":
    raise SystemExit(main())
